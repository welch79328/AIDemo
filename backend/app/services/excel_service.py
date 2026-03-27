"""
Excel 處理服務
用於 Excel 檔案解析、生成與資料驗證
"""
import re
import json
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional
from datetime import datetime
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, or_
from app.models.base import Customer, CustomerEvaluation, AIAnalysis


class ExcelService:
    """Excel 處理服務類"""

    # 類別常數：驗證規則
    MOBILE_PATTERN = re.compile(r'^09\d{8}$')  # 台灣手機：09 開頭 10 碼
    LANDLINE_PATTERN = re.compile(r'^0[2-8]\d{7,8}$')  # 台灣市話：02-08 開頭 9-10 碼
    EMAIL_PATTERN = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')

    # 必填欄位清單（至少需要聯絡人和聯絡電話）
    REQUIRED_FIELDS = ["contact_person", "contact_phone"]

    # 標題映射（Excel 標題 -> 資料庫欄位）
    HEADER_MAPPING = {
        # 原有映射（通用格式）
        "公司名稱": "company_name",
        "聯絡人": "contact_person",
        "聯絡電話": "contact_phone",
        "Email": "contact_email",
        "電子郵件": "contact_email",
        "地址": "address",
        "廣告來源": "ad_source",

        # 3特點輪播廣告 專用映射
        "個人姓名": "contact_person",
        "電話(手機)": "contact_phone",
        "信箱": "contact_email",
        "目前房屋所在縣市": "address",
        "從哪一個廣告來源？\n（後台分析）": "ad_source",
        # 注意：職業別、管理房源數量、負責業務等欄位暫時不導入
        # 因為這些資料需要另外的資料結構或處理邏輯
    }

    def __init__(self):
        """初始化 ExcelService"""
        pass

    async def parse_lead_import_file(
        self,
        file_path: Path
    ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        解析導入的 Excel 檔案

        Args:
            file_path: Excel 檔案路徑

        Returns:
            (valid_rows, error_rows) - 有效資料和錯誤資料
        """
        valid_rows = []
        error_rows = []

        try:
            # 讀取 Excel 檔案
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active

            # 獲取標題列（第一列）
            headers = self._get_headers(ws)

            # 處理資料列（從第二列開始）
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # 如果整列都是空值，跳過
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    continue

                # 將行資料轉換為字典
                row_data = self._row_to_dict(headers, row)

                # 驗證資料
                is_valid, errors = self._validate_row(row_data, row_idx)

                if is_valid:
                    valid_rows.append(row_data)
                else:
                    # 記錄錯誤
                    for error in errors:
                        error_rows.append({
                            "row_number": row_idx,
                            "error_type": error["type"],
                            "error_message": error["message"],
                            "row_data": row_data
                        })

            wb.close()

        except Exception as e:
            raise Exception(f"Excel 檔案解析失敗: {str(e)}")

        return valid_rows, error_rows

    def _get_headers(self, ws: Worksheet) -> List[str]:
        """
        獲取標題列並轉換為標準欄位名稱

        Args:
            ws: Worksheet 物件

        Returns:
            標準化的欄位名稱列表
        """
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

        headers = []
        for cell in header_row:
            cell_value = str(cell).strip() if cell else ""
            # 轉換為標準欄位名稱
            standard_name = self.HEADER_MAPPING.get(cell_value, cell_value)
            headers.append(standard_name)

        return headers

    def _row_to_dict(self, headers: List[str], row: tuple) -> Dict[str, Any]:
        """
        將行資料轉換為字典

        Args:
            headers: 標題列表
            row: 行資料 tuple

        Returns:
            資料字典
        """
        row_dict = {}
        for i, header in enumerate(headers):
            if i < len(row):
                value = row[i]
                # 轉換為字串並去除前後空白
                row_dict[header] = str(value).strip() if value is not None else ""
            else:
                row_dict[header] = ""

        # 清理電話號碼格式
        if "contact_phone" in row_dict and row_dict["contact_phone"]:
            row_dict["contact_phone"] = self._clean_phone_number(row_dict["contact_phone"])

        return row_dict

    def _clean_phone_number(self, phone: str) -> str:
        """
        清理電話號碼格式
        - 移除 Excel 公式前綴 = 和 p:
        - 移除國碼 +886 或 886
        - 轉換為台灣手機格式 09XXXXXXXX 或市話格式 0X-XXXX-XXXX

        Args:
            phone: 原始電話號碼

        Returns:
            清理後的電話號碼
        """
        if not phone:
            return ""

        # 移除空白
        phone = phone.strip()

        # 移除 Excel 公式前綴 =
        if phone.startswith("="):
            phone = phone[1:]

        # 移除 Excel p: 前綴（某些版本會加上）
        if phone.startswith("p:"):
            phone = phone[2:]

        # 移除 + 符號
        phone = phone.replace("+", "")

        # 移除破折號、空格、括號
        phone = phone.replace("-", "").replace(" ", "").replace("(", "").replace(")", "")

        # 處理國碼 886
        if phone.startswith("886"):
            phone = phone[3:]  # 移除 886
            # 如果移除後是 9 開頭（手機），加上 0
            if phone and phone[0] == "9":
                phone = "0" + phone
            # 如果移除後是 2-8 開頭（市話區碼），加上 0
            elif phone and phone[0] in "2345678":
                phone = "0" + phone

        # 如果只有 9 碼且是 9 開頭（手機），加上 0
        if len(phone) == 9 and phone[0] == "9":
            phone = "0" + phone

        # 如果只有 8-9 碼且是 2-8 開頭（市話），加上 0
        if len(phone) in [8, 9] and phone and phone[0] in "2345678":
            phone = "0" + phone

        return phone

    def _validate_row(
        self,
        row_data: Dict[str, Any],
        row_number: int
    ) -> Tuple[bool, List[Dict[str, str]]]:
        """
        驗證行資料

        Args:
            row_data: 行資料字典
            row_number: 行號

        Returns:
            (is_valid, errors) - 是否有效和錯誤列表
        """
        errors = []

        # 必填欄位檢查
        for field in self.REQUIRED_FIELDS:
            if not row_data.get(field) or row_data.get(field) == "":
                errors.append({
                    "type": "missing_required_field",
                    "message": f"缺少必填欄位: {field}"
                })

        # 電話格式驗證
        if row_data.get("contact_phone"):
            if not self._validate_phone(row_data["contact_phone"]):
                errors.append({
                    "type": "invalid_phone_format",
                    "message": f"電話格式錯誤: {row_data['contact_phone']}"
                })

        # Email 格式驗證（如果有提供）
        if row_data.get("contact_email") and row_data["contact_email"] != "":
            if not self._validate_email(row_data["contact_email"]):
                errors.append({
                    "type": "invalid_email_format",
                    "message": f"Email 格式錯誤: {row_data['contact_email']}"
                })

        is_valid = len(errors) == 0
        return is_valid, errors

    def _validate_phone(self, phone: str) -> bool:
        """
        驗證電話格式（台灣手機或市話）

        Args:
            phone: 電話號碼

        Returns:
            是否有效
        """
        if not phone:
            return False

        phone = phone.strip()

        # 檢查手機格式或市話格式
        return bool(
            self.MOBILE_PATTERN.match(phone) or
            self.LANDLINE_PATTERN.match(phone)
        )

    def _validate_email(self, email: str) -> bool:
        """
        驗證 Email 格式

        Args:
            email: Email 地址

        Returns:
            是否有效
        """
        if not email:
            return False

        email = email.strip()
        return bool(self.EMAIL_PATTERN.match(email))

    async def detect_duplicates(
        self,
        db: AsyncSession,
        rows: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """
        檢測重複資料（比對資料庫現有客戶）

        Args:
            db: 資料庫 session
            rows: 待檢測的資料列表

        Returns:
            重複資料列表（包含 row_number, customer_name, contact_phone, email, existing_customer_id）
        """
        duplicates = []

        if not rows:
            return duplicates

        # 收集所有電話號碼和 Email（用於批次查詢）
        phones = []
        emails = []

        for row in rows:
            phone = row.get("contact_phone", "").strip()
            email = row.get("contact_email", "").strip()

            if phone:
                phones.append(phone)

            # 只加入非空的 Email
            if email and email != "":
                emails.append(email)

        # 批次查詢資料庫中已存在的客戶
        # 使用 OR 條件：contact_phone IN (...) OR email IN (...)
        query_conditions = []

        if phones:
            query_conditions.append(Customer.contact_phone.in_(phones))

        if emails:
            query_conditions.append(Customer.contact_email.in_(emails))

        if not query_conditions:
            return duplicates

        # 執行查詢
        stmt = select(Customer).where(or_(*query_conditions))
        result = await db.execute(stmt)
        existing_customers = result.scalars().all()

        # 建立快速查找字典（電話 -> Customer, Email -> Customer）
        phone_map = {}
        email_map = {}

        for customer in existing_customers:
            if customer.contact_phone:
                phone_map[customer.contact_phone] = customer

            if customer.contact_email:
                email_map[customer.contact_email] = customer

        # 檢查每一列是否重複
        for idx, row in enumerate(rows, start=1):
            phone = row.get("contact_phone", "").strip()
            email = row.get("contact_email", "").strip()

            # 檢查電話重複
            if phone and phone in phone_map:
                existing_customer = phone_map[phone]
                duplicates.append({
                    "row_number": idx,
                    "customer_name": row.get("company_name", ""),
                    "contact_phone": phone,
                    "contact_email": row.get("contact_email", ""),
                    "existing_customer_id": existing_customer.id
                })
                continue  # 已找到重複，跳到下一列

            # 檢查 Email 重複（只檢查非空 Email）
            if email and email != "" and email in email_map:
                existing_customer = email_map[email]
                duplicates.append({
                    "row_number": idx,
                    "customer_name": row.get("company_name", ""),
                    "contact_phone": row.get("contact_phone", ""),
                    "contact_email": email,
                    "existing_customer_id": existing_customer.id
                })

        return duplicates

    async def generate_health_check_report(
        self,
        customer: Customer,
        evaluation: CustomerEvaluation,
        ai_analysis: Optional[AIAnalysis] = None,
        template_path: Optional[Path] = None
    ) -> Path:
        """
        生成客戶健檢報告 Excel 檔案

        Args:
            customer: 客戶資料
            evaluation: 客戶評估記錄
            ai_analysis: AI 分析結果（可選）
            template_path: 範本檔案路徑（可選）

        Returns:
            生成的檔案路徑
        """
        # 1. 建立新的 Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "健檢報告"

        # 2. 載入問卷資料
        questionnaire_path = Path(__file__).parent.parent / "data" / "questionnaire_30.json"
        with open(questionnaire_path, "r", encoding="utf-8") as f:
            questionnaire_data = json.load(f)

        # 3. 建立報告標題
        ws.append(["客戶健檢報告"])
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:D1")

        ws.append([])  # 空行

        # 4. 客戶基本資料區
        ws.append(["客戶基本資料"])
        ws[f"A{ws.max_row}"].font = Font(size=14, bold=True)

        ws.append(["公司名稱", customer.company_name or ""])
        ws.append(["聯絡人", customer.contact_name or ""])
        ws.append(["聯絡電話", customer.contact_phone or ""])
        ws.append(["Email", customer.contact_email or ""])
        ws.append(["地址", customer.address or ""])
        ws.append(["廣告來源", customer.ad_source or ""])

        ws.append([])  # 空行

        # 5. 評估結果區
        ws.append(["評估結果"])
        ws[f"A{ws.max_row}"].font = Font(size=14, bold=True)

        ws.append(["客戶等級", evaluation.grade.value if evaluation.grade else ""])
        ws.append(["綜合評分", f"{evaluation.score} 分" if evaluation.score else ""])

        # 如果有 AA 客戶判定理由
        if ai_analysis and ai_analysis.is_aa_customer:
            reasons = ", ".join(ai_analysis.aa_reasons) if ai_analysis.aa_reasons else ""
            ws.append(["AA 客戶判定理由", reasons])

        # 覆蓋率
        if ai_analysis and ai_analysis.coverage_rate:
            matched_count = len(ai_analysis.matched_questions) if ai_analysis.matched_questions else 0
            ws.append(["問卷覆蓋率", f"{matched_count}/30 ({ai_analysis.coverage_rate:.1f}%)"])

        ws.append([])  # 空行

        # 6. 業務30問問答區
        ws.append(["業務30問問答記錄"])
        ws[f"A{ws.max_row}"].font = Font(size=14, bold=True)

        ws.append(["問題編號", "階段", "問題內容", "客戶回答", "狀態"])
        header_row = ws.max_row

        # 設定標題列樣式
        for col in range(1, 6):
            cell = ws.cell(row=header_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 建立匹配問題的快速查找字典
        matched_questions_map = {}
        if ai_analysis and ai_analysis.matched_questions:
            for mq in ai_analysis.matched_questions:
                matched_questions_map[mq.get("number")] = mq

        # 7. 填充30問資料
        for q in questionnaire_data:
            q_number = q.get("number")
            matched_q = matched_questions_map.get(q_number)

            if matched_q:
                # 已討論的問題 - 綠色標記
                answer = matched_q.get("answer", "")
                status = "✓ 已討論"
                row_data = [
                    f"Q{q_number}",
                    q.get("phase", ""),
                    q.get("question", ""),
                    answer,
                    status
                ]
                ws.append(row_data)

                # 設定綠色背景
                current_row = ws.max_row
                for col in range(1, 6):
                    cell = ws.cell(row=current_row, column=col)
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                # 未討論的問題 - 黃色標記
                status = "✗ 未討論"
                row_data = [
                    f"Q{q_number}",
                    q.get("phase", ""),
                    q.get("question", ""),
                    "",
                    status
                ]
                ws.append(row_data)

                # 設定黃色背景
                current_row = ws.max_row
                for col in range(1, 6):
                    cell = ws.cell(row=current_row, column=col)
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        # 8. 調整欄寬
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 12

        # 9. 儲存檔案
        # 建立儲存目錄（按日期分組）
        today = datetime.now().strftime("%Y-%m-%d")
        report_dir = Path("storage/reports") / today
        report_dir.mkdir(parents=True, exist_ok=True)

        # 檔案命名：{customer_name}_健檢報告_{date}.xlsx
        customer_name = customer.company_name or customer.contact_name or "未命名客戶"
        # 清理檔案名稱中的特殊字元
        safe_name = re.sub(r'[\\/*?:"<>|]', '_', customer_name)
        filename = f"{safe_name}_健檢報告_{today}.xlsx"
        file_path = report_dir / filename

        wb.save(file_path)
        wb.close()

        return file_path


# 建立全域實例
excel_service = ExcelService()
